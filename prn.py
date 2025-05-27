import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

directorio_prn = os.path.join(os.path.dirname(__file__), 'prn')

if not os.path.isdir(directorio_prn):
    print("❌ La dirección especificada no es válida.")
    exit()

archivos_prn = [f for f in os.listdir(directorio_prn) if f.endswith('.prn')]

if not archivos_prn:
    print("❌ No se encontraron archivos .prn en el directorio especificado.")
    exit()

for archivo_prn in archivos_prn:
    ruta_prn = os.path.join(directorio_prn, archivo_prn)

    with open(ruta_prn, 'r', encoding='utf-8') as f:
        lineas = f.readlines()

    primeras_filas = [line.strip() for line in lineas[:6]]
    primeras_filas = pd.DataFrame([[line] for line in primeras_filas])

    resto = [line.strip() for line in lineas[6:]]
    resto_separado = [line.split(',') for line in resto]
    df_resto = pd.DataFrame(resto_separado)

    for i in range(len(df_resto)):
        fila = df_resto.iloc[i].tolist()
        if len(fila) > 6:
            fila_modificada = fila[:5] + fila[7:]
            while len(fila_modificada) < len(fila):
                fila_modificada.append("")
            df_resto.iloc[i] = fila_modificada

    nombres_columnas = [
        "fecha", "hora", "elemento", "identificacion", "concentracion",
        "rsd", "mean abs", "read 1", "read 2", "read 3"
    ]
    num_columnas = df_resto.shape[1]
    fila_numerada = [nombres_columnas[i] if i < len(nombres_columnas) else "" for i in range(num_columnas)]
    fila_numerada = pd.DataFrame([fila_numerada])

    resultado = pd.concat([primeras_filas, fila_numerada, df_resto], ignore_index=True)

    nombre_base = os.path.splitext(os.path.basename(archivo_prn))[0]
    archivo_salida = os.path.join(directorio_prn, f'{nombre_base}.xlsx')

    resultado.to_excel(archivo_salida, index=False, header=False)

    wb = load_workbook(archivo_salida)
    ws = wb.active

    max_col_index = 0
    for row in ws.iter_rows(min_row=7):
        last_used = max((cell.column for cell in row if cell.value not in (None, "")), default=0)
        max_col_index = max(max_col_index, last_used)

    for col_idx in range(1, max_col_index + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_row=7, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(archivo_salida)

    print(f"✅ ¡Archivo Excel creado exitosamente como '{archivo_salida}' con columnas ajustadas!")

print("✅ ¡Todos los archivos .prn han sido procesados y exportados a Excel!")
